"""Update MP-02 through MP-06 Excel templates with live-validated signal mappings."""
import openpyxl
import copy
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE = r'C:\local\opencode\codesys\templates'

# ---------------------------------------------------------------------------
# Mapping data per template
# Each entry: (new_signal_name, new_description, new_hw_source, note)
# Rows are in the same order as the original data rows (row 3..N).
# ---------------------------------------------------------------------------

MP02 = [
    # (signal_name, description, hw_source, note)
    ('current_travelA', 'DC bus current of travel drive A', 'CANBusDrive/cTransA/Current', 'present'),
    ('current_travelB', 'DC bus current of travel drive B', 'CANBusDrive/cTransB/Current', 'present'),
    ('current_travelC', 'DC bus current of travel drive C', 'CANBusDrive/cTransC/Current', 'present'),
    ('current_travelD', 'DC bus current of travel drive D', 'CANBusDrive/cTransD/Current', 'present'),
    ('batt_voltage_travelA', 'Battery voltage of travel drive A', 'CANBusDrive/cTransA/BattVoltage', 'present'),
    ('travel_rpm_M1', 'Motor RPM drive motor 1', 'N/A', 'no present'),
    ('travel_rpm_M2', 'Motor RPM drive motor 2', 'N/A', 'no present'),
    ('travel_rpm_M3', 'Motor RPM drive motor 3', 'N/A', 'no present'),
    ('travel_rpm_M4', 'Motor RPM drive motor 4', 'N/A', 'no present'),
    ('torque_travelA', 'Estimated motor torque of travel drive A', 'CANBusDrive/cTransA/Torque', 'present'),
    ('vehicle_speed_encoder', 'Vehicle speed from wheel encoder', 'N/A', 'no present'),
    ('vehicle_speed_gps', 'Vehicle speed from GPS RTK (external)', 'N/A', 'manual-only'),
    ('position_steerA', 'Steering position of wheel A (front-left)', 'System/iSteerAPos', 'present'),
    ('position_steerB', 'Steering position of wheel B (front-right)', 'System/iSteerBPos', 'present'),
    ('position_steerC', 'Steering position of wheel C (rear-left)', 'System/iSteerCPos', 'present'),
    ('position_steerD', 'Steering position of wheel D (rear-right)', 'System/iSteerDPos', 'present'),
    ('motor_temp_steerA', 'Motor winding temperature of steer drive A', 'CANBusDrive/cSteerA/MotorTemp', 'present'),
    ('joystick_cmd_x', 'Joystick command X axis (forward/back)', 'N/A', 'manual-only'),
    ('joystick_cmd_y', 'Joystick command Y axis (left/right)', 'N/A', 'manual-only'),
    ('bms_current', 'Battery management system current (signed)', 'System/BMSAB/Current', 'present'),
    ('bms_voltage', 'Battery management system voltage', 'System/BMSAB/Voltage', 'present'),
    ('travel_fault_code_any', 'Drive controller fault code (any drive)', 'N/A', 'no present'),
    ('operating_mode', 'Vehicle operating mode state', 'System/OperatingMode', 'not validated'),
    ('distance_m', 'Odometer distance this run', 'N/A', 'no present'),
]

MP03 = [
    ('bms_voltage', 'Battery pack terminal voltage', 'System/BMSAB/Voltage', 'present'),
    ('bms_current', 'Battery pack current (signed, +discharge / -regen)', 'System/BMSAB/Current', 'present'),
    ('bms_soc', 'Battery state of charge', 'System/BMSAB/SOC', 'present'),
    ('bat_soh_pct', 'Battery state of health', 'N/A', 'no present'),
    ('bat_energy_discharged_kwh', 'Cumulative energy discharged this session', 'N/A', 'no present'),
    ('bat_energy_regen_kwh', 'Cumulative regen energy recovered this session', 'N/A', 'no present'),
    ('bat_temp_max', 'Maximum cell temperature across all modules', 'N/A', 'no present'),
    ('bat_temp_min', 'Minimum cell temperature across all modules', 'N/A', 'no present'),
    ('bat_cell_vmin', 'Minimum individual cell voltage', 'System/BMSAB', 'not validated'),
    ('bat_cell_vmax', 'Maximum individual cell voltage', 'System/BMSAB', 'not validated'),
    ('bat_fault_code', 'BMS fault code', 'N/A', 'no present'),
    ('bat_cycle_count', 'Total charge-discharge cycle count', 'System/BMSAB', 'not validated'),
    ('bat_charge_power_kw', 'Charger power input during charging session', 'System/BMSAB', 'not validated'),
    ('ambient_temp', 'Ambient temperature', 'N/A', 'manual-only'),
]

MP04 = [
    ('motor_temp_travelA', 'Travel motor A winding temperature', 'CANBusDrive/cTransA/MotorTemp', 'present'),
    ('motor_temp_travelB', 'Travel motor B winding temperature', 'CANBusDrive/cTransB/MotorTemp', 'present'),
    ('motor_temp_travelC', 'Travel motor C winding temperature', 'CANBusDrive/cTransC/MotorTemp', 'present'),
    ('motor_temp_travelD', 'Travel motor D winding temperature', 'CANBusDrive/cTransD/MotorTemp', 'present'),
    ('cntrl_temp_travelA', 'Travel controller A heatsink temperature', 'CANBusDrive/cTransA/CntrlTemp', 'present'),
    ('cntrl_temp_travelB', 'Travel controller B heatsink temperature', 'CANBusDrive/cTransB', 'not validated'),
    ('cntrl_temp_travelC', 'Travel controller C heatsink temperature', 'CANBusDrive/cTransC', 'not validated'),
    ('cntrl_temp_travelD', 'Travel controller D heatsink temperature', 'CANBusDrive/cTransD', 'not validated'),
    ('motor_temp_winchA', 'Hoist motor A winding temperature', 'CANBusDrive/cWinchA/MotorTemp', 'present'),
    ('motor_temp_winchB', 'Hoist motor B winding temperature', 'CANBusDrive/cWinchB/MotorTemp', 'present'),
    ('motor_temp_winchC', 'Hoist motor C winding temperature', 'CANBusDrive/cWinchC/MotorTemp', 'present'),
    ('motor_temp_winchD', 'Hoist motor D winding temperature', 'CANBusDrive/cWinchD/MotorTemp', 'present'),
    ('steer_temp_motor_max', 'Maximum steer motor temperature (all 4)', 'CANBusDrive/cSteerA', 'not validated'),
    ('vibration_motor_M1', 'Vibration RMS travel motor 1', 'N/A', 'manual-only'),
    ('vibration_motor_M2', 'Vibration RMS travel motor 2', 'N/A', 'manual-only'),
    ('hydraulic_oil_temp', 'Hydraulic oil temperature', 'System/Hydraulic', 'not validated'),
    ('hydraulic_pressure', 'Hydraulic pump output pressure', 'System/Hydraulic', 'not validated'),
    ('insulation_resistance', 'Motor insulation resistance', 'N/A', 'manual-only'),
    ('ambient_temp', 'Ambient temperature', 'N/A', 'manual-only'),
]

MP05 = [
    ('bat_soh_pct', 'Battery state of health', 'N/A', 'no present'),
    ('bat_cycle_count', 'Total charge-discharge cycle count', 'System/BMSAB', 'not validated'),
    ('bat_capacity_ah', 'Measured usable battery capacity this cycle', 'System/BMSAB', 'not validated'),
    ('bat_r_internal_mohm', 'Battery internal resistance', 'System/BMSAB', 'not validated'),
    ('encoder_linearity_pct', 'Encoder linearity check (all axes)', 'System/Encoder', 'not validated'),
    ('fault_code_count_total', 'Total fault events this session', 'System/PLC', 'not validated'),
    ('fault_code_log', 'Last fault code and timestamp', 'System/PLC', 'not validated'),
    ('radio_signal_loss_pct', 'Radio remote signal loss rate', 'System/Radio', 'not validated'),
    ('radio_response_ms', 'Radio command response time', 'System/Radio', 'not validated'),
    ('tire_pressure_FL', 'Tire pressure front-left', 'N/A', 'manual-only'),
    ('tire_pressure_FR', 'Tire pressure front-right', 'N/A', 'manual-only'),
    ('tire_pressure_RL', 'Tire pressure rear-left', 'N/A', 'manual-only'),
    ('tire_pressure_RR', 'Tire pressure rear-right', 'N/A', 'manual-only'),
    ('brake_pad_thickness_mm', 'Brake pad thickness measurement', 'N/A', 'manual-only'),
    ('hydraulic_filter_dp_bar', 'Hydraulic filter differential pressure', 'System/Hydraulic', 'not validated'),
    ('connector_resistance_mohm', 'Connector and cable resistance spot check', 'N/A', 'manual-only'),
    ('hmi_fault_export_count', 'HMI fault log entry count (session)', 'N/A', 'manual-only'),
    ('runtime_hours_total', 'Total vehicle runtime hours', 'System/Vehicle', 'not validated'),
]

MP06 = [
    ('batt_voltage_travelA', 'DC bus voltage of travel drive A', 'CANBusDrive/cTransA/BattVoltage', 'present'),
    ('current_travelA', 'DC bus current of travel drive A (signed)', 'CANBusDrive/cTransA/Current', 'present'),
    ('batt_voltage_travelB', 'DC bus voltage of travel drive B', 'CANBusDrive/cTransB', 'not validated'),
    ('current_travelB', 'DC bus current of travel drive B (signed)', 'CANBusDrive/cTransB/Current', 'present'),
    ('batt_voltage_travelC', 'DC bus voltage of travel drive C', 'CANBusDrive/cTransC', 'not validated'),
    ('current_travelC', 'DC bus current of travel drive C (signed)', 'CANBusDrive/cTransC/Current', 'present'),
    ('batt_voltage_travelD', 'DC bus voltage of travel drive D', 'CANBusDrive/cTransD', 'not validated'),
    ('current_travelD', 'DC bus current of travel drive D (signed)', 'CANBusDrive/cTransD/Current', 'present'),
    ('torque_travelA', 'Motor torque of travel drive A (estimated)', 'CANBusDrive/cTransA/Torque', 'present'),
    ('torque_travelB', 'Motor torque of travel drive B (estimated)', 'CANBusDrive/cTransB', 'not validated'),
    ('torque_travelC', 'Motor torque of travel drive C (estimated)', 'CANBusDrive/cTransC', 'not validated'),
    ('torque_travelD', 'Motor torque of travel drive D (estimated)', 'CANBusDrive/cTransD', 'not validated'),
    ('travel_rpm_M1', 'Motor RPM of travel drive A', 'N/A', 'no present'),
    ('travel_rpm_M2', 'Motor RPM of travel drive B', 'N/A', 'no present'),
    ('travel_rpm_M3', 'Motor RPM of travel drive C', 'N/A', 'no present'),
    ('travel_rpm_M4', 'Motor RPM of travel drive D', 'N/A', 'no present'),
    ('bms_voltage', 'Battery pack voltage', 'System/BMSAB/Voltage', 'present'),
    ('bms_current', 'Battery pack current (signed)', 'System/BMSAB/Current', 'present'),
    ('vehicle_speed_gps', 'Vehicle speed from GPS RTK (ground truth)', 'N/A', 'manual-only'),
    ('vehicle_speed_encoder', 'Vehicle speed from wheel encoder', 'N/A', 'no present'),
    ('operating_mode', 'Vehicle operating mode state', 'System/OperatingMode', 'not validated'),
    ('hoist_load_kg', 'Container load on hoist', 'System/Hoist', 'not validated'),
]

TEMPLATES = {
    'MP-02.xlsx': MP02,
    'MP-03.xlsx': MP03,
    'MP-04.xlsx': MP04,
    'MP-05.xlsx': MP05,
    'MP-06.xlsx': MP06,
}

def update_template(filename, mapping):
    """Update a single template file with the given mapping."""
    filepath = f'{BASE}\\{filename}'
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Check if Note column exists
    headers = [c.value for c in ws[1]]
    has_note = 'Note' in headers
    note_col = None
    if has_note:
        note_col = headers.index('Note') + 1  # 1-based
    else:
        # Add Note as column J (col 10)
        note_col = 10
        ws.cell(row=2, column=note_col, value='Note')

    # Update each data row
    for i, (sig_name, desc, hw_src, note) in enumerate(mapping):
        row = 3 + i  # data starts at row 3
        if row > ws.max_row:
            print(f'  WARNING: row {row} exceeds max row {ws.max_row} in {filename}')
            continue

        ws.cell(row=row, column=1, value=sig_name)       # Signal Name
        ws.cell(row=row, column=2, value=desc)            # Description
        ws.cell(row=row, column=4, value=hw_src)          # Hardware Source
        ws.cell(row=row, column=note_col, value=note)     # Note

    wb.save(filepath)
    print(f'  Saved {filename}')

    # Count notes
    note_counts = {}
    for i in range(len(mapping)):
        row = 3 + i
        val = ws.cell(row=row, column=note_col).value
        note_counts[val] = note_counts.get(val, 0) + 1
    return note_counts

# Process all templates
for filename, mapping in TEMPLATES.items():
    print(f'Processing {filename}...')
    counts = update_template(filename, mapping)
    print(f'  Note counts: {counts}')
    print()

print('Done.')
